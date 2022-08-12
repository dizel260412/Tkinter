from tkinter import *
from tkinter import messagebox
import datetime
from patch import path_to, path_from
from openpyxl import load_workbook

month = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9: 'Сентябрь',
         10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}
k = {'BT': 'RU-Business Trip', 'RW': 'RU-Remote Work', 'CR': 'Ru-Competitor Monitoring', 'CF': 'RU-Child Feeding',
     'ME': 'RU-Medical Examination', 'Trp': 'RU-Training P', 'TB': 'RU-Team Building', 'TrU': 'RU-Training U',
     'OtherP': 'RU-Other P', 'OtherU': 'RU-Other U', 'English': 'RU-English Lessons'}


def click_btn():
    name = txt2.get()
    if name == "":
        name = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    if path_from != txt.get() or path_to != txt1.get():
        a = 'path_from = ' + 'r"{}"'.format(txt.get())
        b = 'path_to = ' + 'r"{}"'.format(txt1.get())
        with open('patch.py', '+w') as file:
            file.write(a + "\n" + b)

    res = f'Путь основного файла: {txt.get()}\nБудет создан файл:      {name}\nПуть созданного файла: ' \
          f'{txt1.get()}\n\nДля продолжения нажмите кнопку - "Готово!"'
    lbl5.configure(text=res, bg='grey', fg='white', font=('Arial Bold', 10), justify=LEFT)
    lbl5.place(relx=.15, rely=.6)

    ok = Button(window, text='Готово!', font=('Arial Bold', 10), bg='green', fg='white',
                command=csv)
    ok.place(relx=.8, rely=.8)

    try:
        wb = load_workbook(txt.get(), data_only=True)
        sh = wb['Schedule']
        m_f = month.get(sh.cell(row=2, column=5).value.date().month)
        m = month.get(datetime.datetime.now().month)
        if datetime.datetime.now().month - sh.cell(row=2, column=5).value.date().month == -1:
            bg = 'green'
        else:
            bg = 'red'
        lbl1 = Label(window, text=f'Текущий месяц: {m}\nМесяц в файле: {m_f}', bg=bg, font=('Arial Bold', 12),
                     justify=LEFT)
        lbl1.place(relx=.0, rely=.1)
    except:
        messagebox.showerror('Error', f'В выбранной папке нет файла! \nSchedule Tool to CSV.xlsm')

    return name


def csv():
    wb = load_workbook(txt.get(), data_only=True)
    sh = wb['Schedule']
    def_row = 'ID|Apply Date|Shift Segment Start Time|Shift Segment End Time|' \
              'Shift Segment Type|Cost Centre|Absence Start Time|Absence Name|Absence Amount'
    x = 5
    y = 3
    day_month = sh.cell(row=1, column=36).value
    file_name = f'{txt1.get()}\{click_btn()}'
    with open(fr'{file_name}.csv', 'w', encoding='UTF-8') as file:
        file.write(f'{def_row}\n')
        file.close()
    while x < day_month + 5:
        try:
            for cellObj in sh['A':'A'][2:]:
                val = sh.cell(row=y, column=x).value
                date = sh.cell(row=2, column=x).value.date().strftime('%Y%m%d')
                start = val[:5]
                finish = val[-5:]
                try:
                    comment = sh.cell(row=y, column=x).comment.text
                    if len(comment) > 10:
                        comment = comment.split('\n')
                        comment = comment[-1].strip()
                    if comment not in k.keys():
                        comment = ''
                except:
                    comment = ''
                if len(str(val)) >= 8:
                    if comment == '':
                        with open(fr'{file_name}.csv', 'a', encoding='UTF-8') as file:
                            file.write(f'{cellObj.value}|{date}|{start}|{finish}|Regular||||\n')
                        y += 1
                    else:
                        start_ = datetime.datetime.strptime(start, '%H:%M')
                        finish_ = datetime.datetime.strptime(finish, '%H:%M')
                        amount = str(finish_ - start_)[:-3]
                        with open(fr'{file_name}.csv', 'a', encoding='UTF-8') as file:
                            file.write(f'{cellObj.value}|{date}|||||{start}|{k.get(comment)}|{amount}\n')
                            y += 1
                else:
                    y += 1
        except:
            y = 3
            x += 1


window = Tk()
window.title('CSV_Tool')
window.geometry('1000x600')
# window.configure(bg='blue')

label1 = Label(text='Привет, заполни пожалуйста необходимую информацию.', font=('Comic Sans MS bold', 17),
               fg='green')
label1.pack()

lbl2 = Label(window, text='Путь основного файла:', font=('Arial Bold', 15),
             fg='black', justify=LEFT)
lbl2.place(relx=.0, rely=.3)
txt = Entry(window, width=80)
txt.insert(0, f'{path_from}')
txt.place(relx=.3, rely=.3)

lbl3 = Label(window, text='Путь файла CSV:', font=('Arial Bold', 15),
             fg='black', justify=LEFT)
lbl3.place(relx=.0, rely=.4)
txt1 = Entry(window, width=80)
txt1.insert(0, f'{path_to}')
txt1.place(relx=.3, rely=.4)

lbl4 = Label(window, text='Введите желаемое название файла:', font=('Arial Bold', 15),
             fg='black', justify=LEFT)
lbl4.place(relx=.0, rely=.5)

txt2 = Entry(window, width=40)
txt2.focus()
txt2.place(relx=.4, rely=.5)

lbl5 = Label(window)

btn = Button(window, text='OK!!', font=('Arial Bold', 10), bg='green', fg='white',
             command=click_btn)
btn.place(relx=.8, rely=.5)

window.mainloop()

